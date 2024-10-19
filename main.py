# from datetime import datetime, time, timedelta
# import openpyxl

# woorkbook = openpyxl.load_workbook("relatorio.xlsx")
# sheet = woorkbook["relatorio"]
# data: dict = {"Informações do relatório": [], "Visão Geral": [], "Agentes": []}

# # Informações do relatorio
# inicio = None
# fim = None

# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Informacões do relatório":
#             inicio = cell.row + 1
#         if cell.value == "Visão Geral":
#             fim = cell.row - 1


# for row in sheet.iter_rows(min_row=inicio, max_row=fim):

#     chave = row[0].value
#     valor = row[1].value

#     if isinstance(valor, datetime):
#         valor = valor.strftime("%d/%m/%Y")  # Converte para formato 'DD/MM/AAAA'
#     elif isinstance(valor, time):
#         valor = valor.strftime("%H:%M:%S")  # Converte para formato 'HH:MM:SS'
#     elif isinstance(valor, timedelta):
#         total_seconds = int(valor.total_seconds())
#         hours, remainder = divmod(total_seconds, 3600)
#         minutes, seconds = divmod(remainder, 60)
#         valor = f"{hours}h {minutes}m {seconds}s"
#     if chave is not None and valor is not None:
#         data["Informações do relatório"].append({chave: valor})


# # Visão Geral
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Visão Geral":
#             inicio = cell.row + 1
#         if cell.value == "Total de acesso por agentes":
#             fim = cell.row - 1


# for row in sheet.iter_rows(min_row=inicio, max_row=fim):

#     chave = row[0].value
#     valor = row[1].value

#     if isinstance(valor, datetime):
#         valor = valor.strftime("%d/%m/%Y")  # Converte para formato 'DD/MM/AAAA'
#     elif isinstance(valor, time):
#         valor = valor.strftime("%H:%M:%S")  # Converte para formato 'HH:MM:SS'
#     elif isinstance(valor, timedelta):
#         total_seconds = int(valor.total_seconds())
#         hours, remainder = divmod(total_seconds, 3600)
#         minutes, seconds = divmod(remainder, 60)
#         valor = f"{hours}h {minutes}m {seconds}s"
#     if chave is not None and valor is not None:
#         data["Visão Geral"].append({chave: valor})


# # Pegando o nome de todos agentes
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Total de acesso por agentes":
#             inicio = cell.row + 1
#         if cell.value == "Pesquisa de Satisfação Total da URA":
#             fim = cell.row - 1


# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         if row[0].value.strip() not in [
#             "Agente",
#             "Total de acesso por agentes",
#             "Chamadas atendidas por agentes",
#             "Total de ligações realizadas por agentes",
#             "Total de pausas realizadas por agentes",
#             "Quantidade de classificação por agente",
#             "Pesquisa de Satisfação Total do Operador",
#         ]:
#             data["Agentes"].append(row[0].value)


# data["Agentes"] = list(set(data["Agentes"]))


# # Pegando o dados da planilha Total de acesso
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Total de acesso por agentes":
#             inicio = cell.row + 1
#         if cell.value == "Chamadas atendidas por agentes":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     for i, agente in enumerate(data["Agentes"]):
#         if row[0].value is not None:
#             if data["Agentes"][i] == row[0].value:
#                 data["Agentes"][i] = {
#                     "Agente": agente,
#                     "Qt.Acesso": row[1].value,
#                     "Tempo Total": row[2].value,
#                     "%Tempo Total": row[3].value,
#                 }


# # Chamadas atendidas por agentes
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Chamadas atendidas por agentes":
#             inicio = cell.row + 1
#         if cell.value == "Total de ligações realizadas por agentes":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         for i, agente in enumerate(data["Agentes"]):
#             if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
#                 data["Agentes"][i]["Chamadas Atendidas"] = row[1].value
#                 data["Agentes"][i]["%Chamadas"] = row[2].value
#                 data["Agentes"][i]["Tempo de atendimento"] = row[3].value
#                 data["Agentes"][i]["%Tempo de atendimento"] = row[4].value
#                 data["Agentes"][i]["TMA"] = row[5].value
#                 data["Agentes"][i]["Tempo de espera"] = row[6].value
#                 data["Agentes"][i]["%Tempo de espera"] = row[7].value
#                 data["Agentes"][i]["TME"] = row[8].value

# #Total de ligações realizadas
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Total de ligações realizadas por agentes":
#             inicio = cell.row + 1
#         if cell.value == "Total de pausas realizadas por agentes":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         for i, agente in enumerate(data["Agentes"]):
#             if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
#                 data["Agentes"][i]["Chamadas Realizadas"] = row[1].value
#                 data["Agentes"][i]["%Chamadas"] = row[2].value
#                 data["Agentes"][i]["Tempo das chamadas"] = row[3].value
#                 data["Agentes"][i]["%Tempo das chamadas"] = row[4].value
#                 data["Agentes"][i]["Tempo médio das chamadas"] = row[5].value

# # Pausas
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Total de pausas realizadas por agentes":
#             inicio = cell.row + 1
#         if cell.value == "Quantidade de classificação por agente":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         for i, agente in enumerate(data["Agentes"]):
#             if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
#                 data["Agentes"][i]["Qt. de Pausas"] = row[1].value
#                 data["Agentes"][i]["Tempo Total"] = row[2].value
#                 data["Agentes"][i]["%Tempo Total"] = row[3].value


# # Classificação
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Quantidade de classificação por agente":
#             inicio = cell.row + 1
#         if cell.value == "Pesquisa de Satisfação Total do Operador":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         for i, agente in enumerate(data["Agentes"]):
#             if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
#                 data["Agentes"][i]["Qt. de Classificação"] = row[1].value
#                 data["Agentes"][i]["%Classificação"] = row[2].value





# # NPS Operador
# for row in sheet.iter_rows():
#     for cell in row:
#         if cell.value == "Pesquisa de Satisfação Total do Operador":
#             inicio = cell.row + 1
#         if cell.value == "Pesquisa de Satisfação Total da URA":
#             fim = cell.row - 1

# for row in sheet.iter_rows(min_row=inicio, max_row=fim):
#     if row[0].value is not None:
#         for i, agente in enumerate(data["Agentes"]):
#             if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
#                 data["Agentes"][i]["Qt. de Notas"] = row[1].value
#                 data["Agentes"][i]["Media de Notas"] = row[2].value


# print(data)

from datetime import datetime, time, timedelta
import openpyxl
import json

def formatar_valor(valor):
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")  # Converte para 'DD/MM/AAAA'
    elif isinstance(valor, time):
        return valor.strftime("%H:%M:%S")  # Converte para 'HH:MM:SS'
    elif isinstance(valor, timedelta):
        total_seconds = int(valor.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours}h {minutes}m {seconds}s"
    return valor

def pegar_intervalo_por_valor(sheet, valor_inicio, valor_fim):
    inicio, fim = None, None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == valor_inicio:
                inicio = cell.row + 1
            if cell.value == valor_fim:
                fim = cell.row - 1
    return inicio, fim

def atualizar_agentes(sheet, inicio, fim, data_agentes, campos):
    for row in sheet.iter_rows(min_row=inicio, max_row=fim):
        if row[0].value is not None:
            for agente in data_agentes:
                if isinstance(agente, dict) and agente.get("Agente") == row[0].value:
                    for i, campo in enumerate(campos):
                        agente[campo] = formatar_valor(row[i+1].value)  # Formata o valor aqui

# Carregar planilha
workbook = openpyxl.load_workbook("relatorio.xlsx")
sheet = workbook["relatorio"]

data = {
    "Informações do relatório": [],
    "Visão Geral": [],
    "Agentes": []
}

# Preencher 'Informações do relatório'
inicio, fim = pegar_intervalo_por_valor(sheet, "Informacões do relatório", "Visão Geral")
for row in sheet.iter_rows(min_row=inicio, max_row=fim):
    chave = row[0].value
    valor = formatar_valor(row[1].value)
    if chave is not None and valor is not None:
        data["Informações do relatório"].append({chave: valor})

# Preencher 'Visão Geral'
inicio, fim = pegar_intervalo_por_valor(sheet, "Visão Geral", "Total de acesso por agentes")
for row in sheet.iter_rows(min_row=inicio, max_row=fim):
    chave = row[0].value
    valor = formatar_valor(row[1].value)
    if chave is not None and valor is not None:
        data["Visão Geral"].append({chave: valor})

# Pegando o nome de todos os agentes
inicio, fim = pegar_intervalo_por_valor(sheet, "Total de acesso por agentes", "Pesquisa de Satisfação Total da URA")
for row in sheet.iter_rows(min_row=inicio, max_row=fim):
    if row[0].value is not None:
        agente = row[0].value.strip()
        if agente not in [
            "Agente", "Total de acesso por agentes", "Chamadas atendidas por agentes", 
            "Total de ligações realizadas por agentes", "Total de pausas realizadas por agentes", 
            "Quantidade de classificação por agente", "Pesquisa de Satisfação Total do Operador"
        ]:
            data["Agentes"].append({"Agente": agente})

# Remover duplicatas dos agentes
data["Agentes"] = list({agente['Agente']: agente for agente in data["Agentes"]}.values())

# Atualizar 'Total de acesso por agentes'
inicio, fim = pegar_intervalo_por_valor(sheet, "Total de acesso por agentes", "Chamadas atendidas por agentes")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], ["Qt.Acesso", "Tempo Total de Acesso", "%Tempo Total de Acesso"])

# Atualizar 'Chamadas atendidas por agentes'
inicio, fim = pegar_intervalo_por_valor(sheet, "Chamadas atendidas por agentes", "Total de ligações realizadas por agentes")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], [
    "Chamadas Atendidas", "%Chamadas Atendidas", "Tempo de atendimento das Atendidas", 
    "%Tempo de atendimento das Atendidas", "TMA das Atendidas", "Tempo de espera das Atendidas", "%Tempo de espera das Atendidas", "TME das Atendidas"
])

# Atualizar 'Total de ligações realizadas por agentes'
inicio, fim = pegar_intervalo_por_valor(sheet, "Total de ligações realizadas por agentes", "Total de pausas realizadas por agentes")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], [
    "Chamadas Realizadas", "%Chamadas Realizadas", "Tempo das Chamadas Realizadas", "%Tempo das Chamadas Realizadas", "Tempo médio das chamadas Realizadas"
])

# Atualizar 'Pausas'
inicio, fim = pegar_intervalo_por_valor(sheet, "Total de pausas realizadas por agentes", "Quantidade de classificação por agente")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], ["Qt. de Pausas", "Tempo Total de Pausas", "%Tempo Total de Pausas"])

# Atualizar 'Classificação'
inicio, fim = pegar_intervalo_por_valor(sheet, "Quantidade de classificação por agente", "Pesquisa de Satisfação Total do Operador")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], ["Qt. de Classificação", "%Classificação"])

# Atualizar 'Pesquisa de Satisfação do Operador'
inicio, fim = pegar_intervalo_por_valor(sheet, "Pesquisa de Satisfação Total do Operador", "Pesquisa de Satisfação Total da URA")
atualizar_agentes(sheet, inicio, fim, data["Agentes"], ["Qt. de Notas", "Média de Notas"])

# Imprimir o resultado
with open("dados.json", "w", encoding="utf-8") as ajson:
    json.dump(data, ajson, ensure_ascii=False, indent=4)
    
    
